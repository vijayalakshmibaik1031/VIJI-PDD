from common import *
import importlib.util
import sys


TEST_FILES = [
    "auth_bypass_test.py",
    "rbac_matrix_test.py",
    "idor_test.py",
    "jwt_tampering_test.py",
    "injection_test.py",
    "rate_limit_test.py",
    "headers_test.py",
    "cors_test.py",
    "input_validation_test.py",
    "file_upload_test.py",
    "secret_scan.py",
]


def load_module(path):
    module_name = path.stem
    spec = importlib.util.spec_from_file_location(module_name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def run_test_file(file_name):
    module = load_module(BASE_DIR / file_name)
    if hasattr(module, "run"):
        return module.run()
    return []


def build_excel(report_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except Exception:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = ["endpoint", "method", "role", "status", "expected_status", "finding", "severity", "response_time_ms", "test_category", "note", "timestamp"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in load_report():
        ws.append([row.get(k, "") for k in headers])
    wb.save(report_path)


def summarize(records):
    counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for record in records:
        sev = record.get("severity")
        if sev in counts:
            counts[sev] += 1
    print(f"✓ Endpoints discovered: {len(load_discovered_endpoints())}")
    print(f"✓ Total tests executed: {len(records)}")
    print(f"✓ Total findings: {sum(1 for r in records if r.get('finding'))}")
    print(f"✓ Critical findings: {counts['Critical']}")
    print(f"✓ High findings: {counts['High']}")
    print(f"✓ Medium findings: {counts['Medium']}")
    print(f"✓ Low findings: {counts['Low']}")


def main():
    all_records = []
    for test_file in TEST_FILES:
        try:
            all_records.extend(run_test_file(test_file))
        except Exception as exc:
            all_records.append(build_record(test_file, "RUN", "system", 0, 0, f"Runner error in {test_file}: {exc}", "Medium", 0, "runner", now_iso()))
    write_report(all_records)
    summarize(all_records)
    excel_path = ROOT_DIR / "Vulnerability Test Results" / "Vulnerability_Test_Results_Report.xlsx"
    build_excel(excel_path)
    print(f"Excel written to: {excel_path}")


if __name__ == "__main__":
    main()
